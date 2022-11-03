import pandas as pd
from openpyxl import load_workbook

from main import lista

arquivo_excel_1 = 'base_sorteio.xlsx'

df_nomes = pd.read_excel(arquivo_excel_1, sheet_name='Funcionário')
df_produtos = pd.read_excel(arquivo_excel_1, sheet_name='Produtos')

# Criando a o dataframe com os números sorteados
df_lista = pd.DataFrame({'Ordem':lista})

# Criando a página ordem e incluindo a Lista
book = load_workbook(arquivo_excel_1)
writer = pd.ExcelWriter(arquivo_excel_1, engine='openpyxl')
writer.book = book
df_ordem = df_lista.copy()
df_ordem.to_excel(writer, sheet_name='Ordem', index=False)

writer.save()
writer.close()

# realizando o sorteio
df_sorteio = pd.concat([df_ordem, df_nomes], axis = 1)
df_resultado = pd.merge(df_sorteio, df_produtos, on='Ordem')

# salvando o sorteio na lista
book = load_workbook(arquivo_excel_1)
writer = pd.ExcelWriter(arquivo_excel_1, engine='openpyxl')
writer.book = book
df_ordem = df_resultado.copy()
df_ordem.to_excel(writer, sheet_name='Sorteio', index=False)

writer.save()
writer.close()

print(df_resultado)


